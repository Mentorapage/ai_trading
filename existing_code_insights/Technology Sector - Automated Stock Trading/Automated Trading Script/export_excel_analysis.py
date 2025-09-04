"""
EXCEL ANALYSIS EXPORT
=====================
Generate comprehensive Excel workbook with detailed trading analysis including volume metrics
"""

import argparse
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import logging
import os
from typing import Dict, List, Tuple, Any
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.formatting.rule import ColorScaleRule

# Import our modules
from config_loader import config
from trading_core import validate_environment, load_stock_universe, get_sentiment
from trend_filter import apply_trend_filter, compute_moving_average, get_previous_trading_day, fetch_daily_bars
from news_weighting import apply_news_weighting
import finnhub
from dotenv import load_dotenv
from alpaca.data.historical import StockHistoricalDataClient
from alpaca.data.requests import StockBarsRequest
from alpaca.data.timeframe import TimeFrame

# Load environment variables
load_dotenv(dotenv_path=".env")
finn_api_key = os.getenv("finnhubkey")
api_key = os.getenv("apikey")
secret_key = os.getenv("apisecret")

# Initialize clients
finnhub_client = finnhub.Client(api_key=finn_api_key)
historical_client = StockHistoricalDataClient(api_key, secret_key)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def get_trading_days(start_date: str, end_date: str) -> List[datetime]:
    """Get list of trading days between start and end dates"""
    start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    end_dt = datetime.strptime(end_date, '%Y-%m-%d')
    
    trading_days = []
    current_date = start_dt
    
    while current_date <= end_dt:
        # Skip weekends (Monday = 0, Sunday = 6)
        if current_date.weekday() < 5:
            trading_days.append(current_date)
        current_date += timedelta(days=1)
    
    return trading_days

def get_volume_analysis(symbol: str, analysis_date: datetime) -> Dict[str, Any]:
    """
    Get comprehensive volume analysis for a stock
    
    Returns:
        Dict with volume metrics
    """
    volume_data = {
        'yesterday_volume': None,
        'avg_volume_20d': None,
        'volume_vs_avg': None,
        'volume_above_avg': False,
        'volume_percentile': None,
        'volume_trend': 'Unknown'
    }
    
    try:
        # Get 30 days of data to ensure we have enough for 20-day average
        start_date = analysis_date - timedelta(days=40)
        end_date = analysis_date
        
        # Fetch daily volume data
        daily_data = fetch_daily_bars(symbol, start_date, end_date)
        
        if daily_data is None or len(daily_data) < 21:
            logging.warning(f"Insufficient volume data for {symbol}")
            return volume_data
        
        # Get yesterday's date (previous trading day)
        yesterday = get_previous_trading_day(analysis_date)
        yesterday_date = yesterday.date()
        
        # Check if we have yesterday's data
        if yesterday_date not in daily_data.index:
            logging.warning(f"No volume data for {symbol} on {yesterday_date}")
            return volume_data
        
        # Get yesterday's volume
        yesterday_volume = daily_data.loc[yesterday_date, 'volume']
        volume_data['yesterday_volume'] = int(yesterday_volume)
        
        # Calculate 20-day average volume (excluding current day)
        # Get the last 20 trading days before analysis_date
        volume_series = daily_data['volume'].tail(21)[:-1]  # Exclude current day, get last 20
        
        if len(volume_series) >= 20:
            avg_volume_20d = volume_series.mean()
            volume_data['avg_volume_20d'] = int(avg_volume_20d)
            
            # Calculate volume vs average
            volume_ratio = yesterday_volume / avg_volume_20d
            volume_data['volume_vs_avg'] = volume_ratio
            volume_data['volume_above_avg'] = volume_ratio > 1.0
            
            # Calculate percentile of yesterday's volume vs historical
            volume_percentile = (volume_series < yesterday_volume).sum() / len(volume_series) * 100
            volume_data['volume_percentile'] = volume_percentile
            
            # Determine volume trend
            if volume_ratio >= 1.5:
                volume_data['volume_trend'] = 'Very High'
            elif volume_ratio >= 1.2:
                volume_data['volume_trend'] = 'High'
            elif volume_ratio >= 0.8:
                volume_data['volume_trend'] = 'Normal'
            elif volume_ratio >= 0.5:
                volume_data['volume_trend'] = 'Low'
            else:
                volume_data['volume_trend'] = 'Very Low'
        
        logging.debug(f"Volume analysis for {symbol}: {volume_data}")
        
    except Exception as e:
        logging.error(f"Error in volume analysis for {symbol}: {e}")
    
    return volume_data

def analyze_stock_day_enhanced(symbol: str, analysis_date: datetime, 
                              sentiment_min: float, sentiment_max: float,
                              capital_per_stock: float, stop_loss_pct: float, 
                              take_profit_pct: float) -> Dict[str, Any]:
    """
    Perform enhanced analysis for one stock on one day including volume metrics
    """
    # Set decision time to market open (9:30 AM ET)
    decision_time = analysis_date.replace(hour=9, minute=30)
    date_str = analysis_date.strftime('%Y-%m-%d')
    
    # Initialize result dictionary with all required fields
    result = {
        'Date': date_str,
        'Symbol': symbol,
        'Yesterday_Close': None,
        'MA20': None,
        'Trend_Filter_Result': 'N/A',
        'Yesterday_Volume': None,
        'Avg_Volume_20D': None,
        'Volume_Ratio': None,
        'Volume_Above_Average': 'No',
        'Volume_Trend': 'Unknown',
        'Volume_Percentile': None,
        'Num_News_Articles': 0,
        'Weighted_Sentiment_Score': 0.0,
        'Sentiment_Qualified': 'No',
        'Trend_Qualified': 'No',
        'Overall_Qualified': 'No',
        'Trade_Size': 0,
        'Stop_Loss_Pct': stop_loss_pct,
        'Take_Profit_Pct': take_profit_pct,
        'Final_Trade_Action': 'Skip',
        'Expected_Stop_Loss_Price': None,
        'Expected_Take_Profit_Price': None,
        'Risk_Amount': 0,
        'Reward_Amount': 0,
        'Risk_Reward_Ratio': None
    }
    
    try:
        # 1. Get trend analysis (MA20 and yesterday close)
        yesterday_close, ma20 = compute_moving_average(symbol, analysis_date, 20)
        
        if yesterday_close is not None:
            result['Yesterday_Close'] = round(yesterday_close, 2)
            
            # Calculate expected stop loss and take profit prices
            stop_loss_price = yesterday_close * (1 - stop_loss_pct / 100)
            take_profit_price = yesterday_close * (1 + take_profit_pct / 100)
            result['Expected_Stop_Loss_Price'] = round(stop_loss_price, 2)
            result['Expected_Take_Profit_Price'] = round(take_profit_price, 2)
            
            # Calculate risk and reward amounts
            shares = capital_per_stock / yesterday_close
            risk_amount = shares * (yesterday_close - stop_loss_price)
            reward_amount = shares * (take_profit_price - yesterday_close)
            result['Risk_Amount'] = round(risk_amount, 2)
            result['Reward_Amount'] = round(reward_amount, 2)
            
            if risk_amount > 0:
                result['Risk_Reward_Ratio'] = round(reward_amount / risk_amount, 2)
        
        if ma20 is not None:
            result['MA20'] = round(ma20, 2)
        
        # 2. Apply trend filter
        trend_config = {
            'enabled': True,
            'lookback_days': 20,
            'comparator': 'yesterday_gt_ma'
        }
        
        trend_results = apply_trend_filter([symbol], analysis_date, trend_config)
        trend_passed = trend_results.get(symbol, False)
        result['Trend_Filter_Result'] = 'Pass' if trend_passed else 'Fail'
        result['Trend_Qualified'] = 'Yes' if trend_passed else 'No'
        
        # 3. Get volume analysis
        volume_analysis = get_volume_analysis(symbol, analysis_date)
        result['Yesterday_Volume'] = volume_analysis['yesterday_volume']
        result['Avg_Volume_20D'] = volume_analysis['avg_volume_20d']
        result['Volume_Ratio'] = round(volume_analysis['volume_vs_avg'], 2) if volume_analysis['volume_vs_avg'] else None
        result['Volume_Above_Average'] = 'Yes' if volume_analysis['volume_above_avg'] else 'No'
        result['Volume_Trend'] = volume_analysis['volume_trend']
        result['Volume_Percentile'] = round(volume_analysis['volume_percentile'], 1) if volume_analysis['volume_percentile'] else None
        
        # 4. Get news and sentiment analysis
        try:
            # Fetch news for the date
            all_articles = finnhub_client.company_news(symbol, _from=date_str, to=date_str)
            
            if all_articles:
                # Filter articles by decision time
                valid_articles = []
                for article in all_articles:
                    published_time = datetime.fromtimestamp(article['datetime'])
                    if published_time <= decision_time:
                        valid_articles.append(article)
                
                result['Num_News_Articles'] = len(valid_articles)
                
                if valid_articles:
                    # Get sentiment with source weighting
                    sentiment_score = get_sentiment(symbol, date_str, decision_time)
                    result['Weighted_Sentiment_Score'] = round(sentiment_score, 4)
                else:
                    result['Weighted_Sentiment_Score'] = 0.0
            else:
                result['Num_News_Articles'] = 0
                result['Weighted_Sentiment_Score'] = 0.0
                
        except Exception as e:
            logging.warning(f"Error getting news for {symbol} on {date_str}: {e}")
            result['Num_News_Articles'] = 0
            result['Weighted_Sentiment_Score'] = 0.0
        
        # 5. Determine sentiment qualification
        sentiment_qualified = (sentiment_min <= result['Weighted_Sentiment_Score'] <= sentiment_max)
        result['Sentiment_Qualified'] = 'Yes' if sentiment_qualified else 'No'
        
        # 6. Overall qualification (both trend and sentiment must pass)
        overall_qualified = trend_passed and sentiment_qualified
        result['Overall_Qualified'] = 'Yes' if overall_qualified else 'No'
        
        if overall_qualified:
            result['Trade_Size'] = capital_per_stock
            result['Final_Trade_Action'] = 'Buy'
        else:
            result['Trade_Size'] = 0
            result['Final_Trade_Action'] = 'Skip'
        
        # Add rate limiting
        time.sleep(0.5)
        
    except Exception as e:
        logging.error(f"Error analyzing {symbol} on {date_str}: {e}")
    
    return result

def generate_enhanced_analysis_table(start_date: str, end_date: str, 
                                   sentiment_min: float = 0.1, sentiment_max: float = 0.5,
                                   capital_per_stock: float = 1000000,
                                   stop_loss_pct: float = 5.0, take_profit_pct: float = 5.0) -> pd.DataFrame:
    """
    Generate comprehensive analysis table with volume metrics
    """
    print(f"🔄 Generating enhanced analysis table from {start_date} to {end_date}")
    print(f"📊 Parameters: Sentiment [{sentiment_min:.1f}, {sentiment_max:.1f}], Capital: ${capital_per_stock:,.0f}")
    print(f"🎯 Risk Management: SL={stop_loss_pct}%, TP={take_profit_pct}%")
    
    # Get trading days and stock universe
    trading_days = get_trading_days(start_date, end_date)
    stocks = load_stock_universe()
    
    print(f"📅 Processing {len(trading_days)} trading days")
    print(f"📈 Analyzing {len(stocks)} stocks")
    print(f"🔢 Total combinations: {len(trading_days) * len(stocks)}")
    
    # Collect all analysis results
    all_results = []
    
    for i, analysis_date in enumerate(trading_days):
        date_str = analysis_date.strftime('%Y-%m-%d')
        print(f"\n📅 Processing {date_str} ({i+1}/{len(trading_days)})")
        
        day_results = []
        for j, symbol in enumerate(stocks):
            print(f"   📊 {symbol} ({j+1}/{len(stocks)})", end=' ')
            
            result = analyze_stock_day_enhanced(
                symbol, analysis_date, sentiment_min, sentiment_max, 
                capital_per_stock, stop_loss_pct, take_profit_pct
            )
            
            all_results.append(result)
            day_results.append(result)
            
            # Show quick status
            status = "✅" if result['Overall_Qualified'] == 'Yes' else "❌"
            print(f"{status}")
        
        # Show daily summary
        qualified_count = sum(1 for r in day_results if r['Overall_Qualified'] == 'Yes')
        print(f"   📊 Daily summary: {qualified_count}/{len(stocks)} stocks qualified")
    
    # Convert to DataFrame
    df = pd.DataFrame(all_results)
    
    print(f"\n✅ Enhanced analysis complete: {len(df)} total records")
    return df

def create_excel_workbook(df: pd.DataFrame, filename: str, start_date: str, end_date: str):
    """
    Create comprehensive Excel workbook with multiple sheets and formatting
    """
    print(f"📊 Creating Excel workbook: {filename}")
    
    # Create workbook and remove default sheet
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Sheet 1: Detailed Analysis
    ws1 = wb.create_sheet("Detailed Analysis")
    
    # Add data to sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    
    # Format header row
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Format data rows
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell in row:
            cell.border = border
            if cell.column in [3, 4, 8, 9, 10, 15, 16, 17, 18, 19, 20, 21]:  # Numeric columns
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif cell.column in [5, 7, 11, 13, 14, 15, 19]:  # Center-aligned columns
                cell.alignment = center_alignment
    
    # Auto-adjust column widths
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Add conditional formatting
    # Highlight qualified trades
    qualified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for row in range(2, ws1.max_row + 1):
        if ws1[f'O{row}'].value == 'Yes':  # Overall_Qualified column
            for col in range(1, ws1.max_column + 1):
                ws1.cell(row=row, column=col).fill = qualified_fill
    
    # Sheet 2: Summary Statistics
    ws2 = wb.create_sheet("Summary Statistics")
    
    # Calculate summary statistics
    total_records = len(df)
    qualified_records = len(df[df['Overall_Qualified'] == 'Yes'])
    trend_pass = len(df[df['Trend_Qualified'] == 'Yes'])
    sentiment_pass = len(df[df['Sentiment_Qualified'] == 'Yes'])
    volume_above_avg = len(df[df['Volume_Above_Average'] == 'Yes'])
    
    # Summary data
    summary_data = [
        ['Metric', 'Value', 'Percentage'],
        ['Total Records', total_records, '100.0%'],
        ['Qualified Trades', qualified_records, f'{qualified_records/total_records*100:.1f}%'],
        ['Trend Filter Pass', trend_pass, f'{trend_pass/total_records*100:.1f}%'],
        ['Sentiment Filter Pass', sentiment_pass, f'{sentiment_pass/total_records*100:.1f}%'],
        ['Volume Above Average', volume_above_avg, f'{volume_above_avg/total_records*100:.1f}%'],
        ['', '', ''],
        ['Period', f'{start_date} to {end_date}', ''],
        ['Capital per Trade', '$1,000,000', ''],
        ['Stop Loss', '5%', ''],
        ['Take Profit', '5%', ''],
        ['Sentiment Range', '0.1 to 0.5', ''],
    ]
    
    for row_data in summary_data:
        ws2.append(row_data)
    
    # Format summary sheet
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.border = border
            if cell.column == 2 and isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = center_alignment
    
    # Auto-adjust column widths for summary
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 3: Stock Performance
    ws3 = wb.create_sheet("Stock Performance")
    
    # Stock-wise analysis
    stock_performance = df.groupby('Symbol').agg({
        'Overall_Qualified': lambda x: (x == 'Yes').sum(),
        'Trend_Qualified': lambda x: (x == 'Yes').sum(),
        'Sentiment_Qualified': lambda x: (x == 'Yes').sum(),
        'Volume_Above_Average': lambda x: (x == 'Yes').sum(),
        'Weighted_Sentiment_Score': 'mean',
        'Volume_Ratio': 'mean',
        'Trade_Size': 'sum'
    }).round(4)
    
    stock_performance.columns = ['Qualified_Trades', 'Trend_Pass', 'Sentiment_Pass', 
                               'Volume_Above_Avg', 'Avg_Sentiment', 'Avg_Volume_Ratio', 'Total_Capital']
    stock_performance = stock_performance.reset_index()
    
    # Add to sheet
    for r in dataframe_to_rows(stock_performance, index=False, header=True):
        ws3.append(r)
    
    # Format stock performance sheet
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row:
            cell.border = border
            if cell.column > 1:  # Numeric columns
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = center_alignment
    
    # Auto-adjust column widths
    for column in ws3.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws3.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 4: Daily Performance
    ws4 = wb.create_sheet("Daily Performance")
    
    # Daily analysis
    daily_performance = df.groupby('Date').agg({
        'Overall_Qualified': lambda x: (x == 'Yes').sum(),
        'Trend_Qualified': lambda x: (x == 'Yes').sum(),
        'Sentiment_Qualified': lambda x: (x == 'Yes').sum(),
        'Volume_Above_Average': lambda x: (x == 'Yes').sum(),
        'Weighted_Sentiment_Score': 'mean',
        'Volume_Ratio': 'mean',
        'Trade_Size': 'sum'
    }).round(4)
    
    daily_performance.columns = ['Qualified_Trades', 'Trend_Pass', 'Sentiment_Pass', 
                               'Volume_Above_Avg', 'Avg_Sentiment', 'Avg_Volume_Ratio', 'Total_Capital']
    daily_performance = daily_performance.reset_index()
    
    # Add to sheet
    for r in dataframe_to_rows(daily_performance, index=False, header=True):
        ws4.append(r)
    
    # Format daily performance sheet
    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row):
        for cell in row:
            cell.border = border
            if cell.column > 1:  # Numeric columns
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = center_alignment
    
    # Auto-adjust column widths
    for column in ws4.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws4.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(filename)
    print(f"💾 Excel workbook saved: {filename}")

def print_enhanced_summary_stats(df: pd.DataFrame):
    """Print enhanced summary statistics"""
    print("\n" + "=" * 80)
    print("📊 ENHANCED ANALYSIS SUMMARY STATISTICS")
    print("=" * 80)
    
    total_records = len(df)
    qualified_records = len(df[df['Overall_Qualified'] == 'Yes'])
    
    print(f"📋 Total Records: {total_records:,}")
    print(f"✅ Qualified Trades: {qualified_records:,} ({qualified_records/total_records*100:.1f}%)")
    print(f"❌ Skipped Trades: {total_records-qualified_records:,} ({(total_records-qualified_records)/total_records*100:.1f}%)")
    
    # Filter breakdown
    trend_pass = len(df[df['Trend_Qualified'] == 'Yes'])
    sentiment_pass = len(df[df['Sentiment_Qualified'] == 'Yes'])
    volume_above_avg = len(df[df['Volume_Above_Average'] == 'Yes'])
    
    print(f"\n📈 Filter Results:")
    print(f"   Trend Filter Pass: {trend_pass:,} ({trend_pass/total_records*100:.1f}%)")
    print(f"   Sentiment Filter Pass: {sentiment_pass:,} ({sentiment_pass/total_records*100:.1f}%)")
    print(f"   Volume Above Average: {volume_above_avg:,} ({volume_above_avg/total_records*100:.1f}%)")
    
    # Volume statistics
    volume_ratios = df['Volume_Ratio'].dropna()
    if not volume_ratios.empty:
        print(f"\n📊 Volume Analysis:")
        print(f"   Average Volume Ratio: {volume_ratios.mean():.2f}x")
        print(f"   Median Volume Ratio: {volume_ratios.median():.2f}x")
        print(f"   Max Volume Ratio: {volume_ratios.max():.2f}x")
        print(f"   Min Volume Ratio: {volume_ratios.min():.2f}x")
    
    # Sentiment statistics
    sentiment_scores = df['Weighted_Sentiment_Score']
    print(f"\n📰 Sentiment Statistics:")
    print(f"   Average: {sentiment_scores.mean():.4f}")
    print(f"   Median: {sentiment_scores.median():.4f}")
    print(f"   Min: {sentiment_scores.min():.4f}")
    print(f"   Max: {sentiment_scores.max():.4f}")
    
    # Top performing stocks
    print(f"\n🏆 Top Performing Stocks (Most Qualified Trades):")
    stock_qualified = df[df['Overall_Qualified'] == 'Yes']['Symbol'].value_counts().head(5)
    for symbol, count in stock_qualified.items():
        total_days = len(df[df['Symbol'] == symbol])
        success_rate = count / total_days * 100
        print(f"   {symbol}: {count}/{total_days} qualified trades ({success_rate:.1f}%)")
    
    # Best trading days
    print(f"\n📅 Best Trading Days (Most Qualified Stocks):")
    date_qualified = df[df['Overall_Qualified'] == 'Yes']['Date'].value_counts().head(5)
    for date, count in date_qualified.items():
        total_stocks = len(df[df['Date'] == date])
        success_rate = count / total_stocks * 100
        print(f"   {date}: {count}/{total_stocks} qualified stocks ({success_rate:.1f}%)")
    
    # Total potential capital
    total_capital = qualified_records * 1000000
    print(f"\n💰 Total Capital Deployed: ${total_capital:,.0f}")
    
    # Risk/Reward Analysis
    qualified_df = df[df['Overall_Qualified'] == 'Yes']
    if not qualified_df.empty:
        avg_risk_reward = qualified_df['Risk_Reward_Ratio'].mean()
        total_risk = qualified_df['Risk_Amount'].sum()
        total_reward = qualified_df['Reward_Amount'].sum()
        
        print(f"\n⚖️  Risk/Reward Analysis:")
        print(f"   Average Risk/Reward Ratio: {avg_risk_reward:.2f}")
        print(f"   Total Risk Amount: ${total_risk:,.0f}")
        print(f"   Total Reward Potential: ${total_reward:,.0f}")

def main():
    """Main function"""
    parser = argparse.ArgumentParser(description='Generate comprehensive Excel trading analysis')
    parser.add_argument('--start', required=True, help='Start date (YYYY-MM-DD)')
    parser.add_argument('--end', required=True, help='End date (YYYY-MM-DD)')
    parser.add_argument('--sentiment-min', type=float, default=0.1, help='Minimum sentiment threshold')
    parser.add_argument('--sentiment-max', type=float, default=0.5, help='Maximum sentiment threshold')
    parser.add_argument('--capital', type=float, default=1000000, help='Capital per stock')
    parser.add_argument('--stop-loss', type=float, default=5.0, help='Stop loss percentage')
    parser.add_argument('--take-profit', type=float, default=5.0, help='Take profit percentage')
    
    args = parser.parse_args()
    
    try:
        # Validate environment
        validate_environment()
        print("✅ Environment validation passed")
        
        # Generate enhanced analysis table
        df = generate_enhanced_analysis_table(
            args.start, args.end, args.sentiment_min, args.sentiment_max, 
            args.capital, args.stop_loss, args.take_profit
        )
        
        # Generate output filename
        start_clean = args.start.replace('-', '_')
        end_clean = args.end.replace('-', '_')
        excel_filename = f"enhanced_analysis_{start_clean}_to_{end_clean}.xlsx"
        
        # Create Excel workbook
        create_excel_workbook(df, excel_filename, args.start, args.end)
        
        # Print summary statistics
        print_enhanced_summary_stats(df)
        
        print(f"\n🎉 Enhanced analysis complete!")
        print(f"📁 Excel file generated: {excel_filename}")
        print(f"📊 Sheets included:")
        print(f"   • Detailed Analysis - Complete trade-by-trade data")
        print(f"   • Summary Statistics - Key performance metrics")
        print(f"   • Stock Performance - Stock-wise analysis")
        print(f"   • Daily Performance - Day-by-day breakdown")
        
    except Exception as e:
        logging.error(f"Analysis failed: {e}")
        print(f"❌ Error: {e}")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())

